#!/usr/bin/env python3
"""
Compare LLM cognitive process graph analysis results across different models or settings.

This script reads multiple LLM analysis JSON files and computes similarity metrics between
their node type counts, average probabilities, and dependency transition matrices.
"""

import json
import argparse
import numpy as np
from pathlib import Path
from typing import Dict, Any, List, Tuple
import itertools
from collections import OrderedDict
import pandas as pd
import os
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import openpyxl


def load_analysis_file(file_path: str) -> Dict[str, Any]:
    """
    Load an LLM analysis JSON file.
    
    Args:
        file_path: Path to the JSON file
        
    Returns:
        Dictionary containing the JSON data
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def extract_metadata_vectors(metadata: Dict[str, Any]) -> Tuple[OrderedDict, Dict[str, Any]]:
    """
    Extract node type counts, avg_prob and dependency matrices from metadata.
    Also extract node types in a consistent order.
    
    Args:
        metadata: Metadata dictionary from an LLM analysis file
        
    Returns:
        Tuple of (ordered_node_types, metrics_dict)
    """
    metrics = metadata.get("average_summary_metrics", {})
    
    # Get all node types and sort them alphabetically for consistent ordering
    node_types = set()
    for key in ["node_type_counts", "node_avg_prob_sum", "dependency"]:
        if key in metrics:
            node_types.update(metrics[key].keys())
    
    # Create an ordered dictionary of node types for consistent indexing
    ordered_node_types = OrderedDict([(node_type, i) for i, node_type in enumerate(sorted(node_types))])
    
    return ordered_node_types, metrics


def create_vectors_and_matrices(ordered_node_types: OrderedDict, metrics: Dict[str, Any]) -> Dict[str, np.ndarray]:
    """
    Create vectors and matrices for node_type_counts, node_avg_prob_sum, and dependency.
    
    Args:
        ordered_node_types: OrderedDict mapping node types to indices
        metrics: Dictionary containing the metrics
        
    Returns:
        Dictionary of numpy arrays for each metric
    """
    num_node_types = len(ordered_node_types)
    result = {}
    
    # Create node_type_counts vector
    if "node_type_counts" in metrics:
        counts_vector = np.zeros(num_node_types)
        for node_type, count in metrics["node_type_counts"].items():
            if node_type in ordered_node_types:
                counts_vector[ordered_node_types[node_type]] = count
        result["node_type_counts"] = counts_vector
    
    # Create node_avg_prob_sum vector
    if "node_avg_prob_sum" in metrics:
        prob_vector = np.zeros(num_node_types)
        for node_type, prob in metrics["node_avg_prob_sum"].items():
            if node_type in ordered_node_types and prob is not None:
                prob_vector[ordered_node_types[node_type]] = prob
        result["node_avg_prob_sum"] = prob_vector
    
    # Create dependency matrix
    if "dependency" in metrics:
        dep_matrix = np.zeros((num_node_types, num_node_types))
        for from_type, to_types in metrics["dependency"].items():
            if from_type in ordered_node_types:
                for to_type, value in to_types.items():
                    if to_type in ordered_node_types:
                        # Value could be a number or a dict with "num" key
                        # if isinstance(value, dict) and "num" in value:
                        #     dep_matrix[ordered_node_types[from_type], ordered_node_types[to_type]] = value["num"]
                        # else:
                        dep_matrix[ordered_node_types[from_type], ordered_node_types[to_type]] = value
        result["dependency"] = dep_matrix
    
    # Create confidence transition matrix
    # 首先检查node_confidence_transitions是否存在，并记录其结构
    if "all_confidence_transitions" in metrics:
        # 确定哪个键包含信息
#         import pdb; pdb.set_trace()
        transitions_key = "all_confidence_transitions"
        transitions = metrics[transitions_key]
        
        print(f"Found {len(transitions)} transitions in {transitions_key}")
        if transitions and len(transitions) > 0:
            # 显示第一个转换的结构
            print(f"First transition structure: {transitions[0]}")
        
        # Each cell will contain the average confidence value for that transition
        source_prob_matrix = np.zeros((num_node_types, num_node_types))
        target_prob_matrix = np.zeros((num_node_types, num_node_types))
        source_to_target_prob_difference_matrix = np.zeros((num_node_types, num_node_types))
        transition_counts = np.zeros((num_node_types, num_node_types))
        
        for transition in transitions:
            source_type = transition.get("source_node")
            target_type = transition.get("target_node")
            source_prob = transition.get("source_avg_prob")
            target_prob = transition.get("target_avg_prob")
            
            if (source_type in ordered_node_types and target_type in ordered_node_types and 
                source_prob is not None and target_prob is not None):
                source_idx = ordered_node_types[source_type]
                target_idx = ordered_node_types[target_type]
                
                source_prob_matrix[source_idx, target_idx] += source_prob
                target_prob_matrix[source_idx, target_idx] += target_prob
                source_to_target_prob_difference_matrix[source_idx, target_idx] += (source_prob - target_prob)
                transition_counts[source_idx, target_idx] += 1
        
        # Calculate averages only for cells with counts > 0
        nonzero_mask = transition_counts > 0
        result['transition_counts'] = transition_counts
        if np.any(nonzero_mask):  # 确保至少有一个非零元素
            source_prob_matrix[nonzero_mask] /= transition_counts[nonzero_mask]
            target_prob_matrix[nonzero_mask] /= transition_counts[nonzero_mask]
            source_to_target_prob_difference_matrix[nonzero_mask] /= transition_counts[nonzero_mask]
            
            result["source_confidence"] = source_prob_matrix
            result["target_confidence"] = target_prob_matrix
            result["confidence_difference"] = source_to_target_prob_difference_matrix
            print(f"Source confidence matrix shape: {source_prob_matrix.shape}")
            print(f"Target confidence matrix shape: {target_prob_matrix.shape}")
            print(f"Confidence difference matrix shape: {source_to_target_prob_difference_matrix.shape}")
        else:
            print("Warning: No valid transitions found in the data")
    
    return result


def compute_cosine_similarity(v1: np.ndarray, v2: np.ndarray) -> float:
    """
    Compute cosine similarity between two vectors.
    
    Args:
        v1: First vector
        v2: Second vector
        
    Returns:
        Cosine similarity in [0, 1]
    """
    norm1 = np.linalg.norm(v1)
    norm2 = np.linalg.norm(v2)
    
    if norm1 == 0 or norm2 == 0:
        return 0
    
    return np.dot(v1, v2) / (norm1 * norm2)


def compute_matrix_similarity(m1: np.ndarray, m2: np.ndarray) -> float:
    """
    Compute similarity between two matrices using Frobenius inner product.
    
    Args:
        m1: First matrix
        m2: Second matrix
        
    Returns:
        Normalized similarity in [0, 1]
    """
    # Flatten matrices and compute cosine similarity
    v1 = m1.flatten()
    v2 = m2.flatten()
    return compute_cosine_similarity(v1, v2)


model_name_list = [
        "results_temperature_0/deepseek-ai-DeepSeek-R1-Distill-Qwen-7B",
        "results_temperature_0/Qwen-Qwen2.5-7B-Instruct",
        "results_temperature_0/meta-llama-Llama-3.1-8B-Instruct",
        "results_temperature_0/deepseek-ai-DeepSeek-R1-Distill-Llama-8B",

        "results_temperature_0/Qwen-Qwen2.5-3B-Instruct",

        'results_temperature_0/openrouter-qwen-qwq-32b',
        'results_temperature_0/openrouter-microsoft-phi-4'
        
    ]
def get_model_name(file_path: str) -> str:
    if "openrouter-qwen-qwq-32b" in file_path:
        return "openrouter-qwen-qwq-32b"
    elif "openrouter-microsoft-phi-4" in file_path:
        return "openrouter-microsoft-phi-4"
    elif "Qwen-Qwen2.5-3B-Instruct" in file_path:
        return "Qwen-Qwen2.5-3B-Instruct"
    elif "deepseek-ai-DeepSeek-R1-Distill-Llama-8B" in file_path:
        return "deepseek-ai-DeepSeek-R1-Distill-Llama-8B"
    elif "meta-llama-Llama-3.1-8B-Instruct" in file_path:
        return "meta-llama-Llama-3.1-8B-Instruct"
    elif "Qwen-Qwen2.5-7B-Instruct" in file_path:
        return "Qwen-Qwen2.5-7B-Instruct"
    elif "deepseek-ai-DeepSeek-R1-Distill-Qwen-7B" in file_path:
        return "deepseek-ai-DeepSeek-R1-Distill-Qwen-7B"
    else:
        return file_path
    
def create_node_type_counts_chart(file_paths: List[str], all_vectors_matrices: List[Dict[str, np.ndarray]], 
                                 ordered_node_types: OrderedDict, output_dir: Path, normalize: bool = False,
                                 data_key: str = "node_type_counts") -> str:
    """
    Create a bar chart comparing node metrics across all models.
    
    Args:
        file_paths: List of analysis file paths
        all_vectors_matrices: List of vectors and matrices for each file
        ordered_node_types: OrderedDict of node types
        output_dir: Directory to save the chart
        normalize: Whether to normalize counts as proportions
        data_key: The key for the data to plot (node_type_counts or node_avg_prob_sum)
        
    Returns:
        Path to the saved chart file
    """
    # Extract file names for legend
    file_names = [get_model_name(path) for path in file_paths]
    
    # Collect data for each file
    data = {}
    for i, (file_name, vectors_matrices) in enumerate(zip(file_names, all_vectors_matrices)):
        if data_key in vectors_matrices:
            values = vectors_matrices[data_key]
            if normalize and data_key == "node_type_counts":
                # Normalize to sum to 1 (only for counts, not for probabilities)
                total = np.sum(values)
                if total > 0:
                    values = values / total
            data[file_name] = values
    
    if not data:
        print(f"No {data_key} data available for plotting")
        return ""
    
    # Create DataFrame for plotting
    df_data = {}
    node_types = list(ordered_node_types.keys())
    
    for file_name, values in data.items():
        for node_type, idx in ordered_node_types.items():
            if node_type not in df_data:
                df_data[node_type] = []
            df_data[node_type].append(values[idx])
    
    # Create plot
    fig = plt.figure(figsize=(12, 8))
    ax = fig.add_subplot(111)
    
    # Set up bar positions
    num_files = len(data)
    num_types = len(node_types)
    bar_width = 0.8 / num_files
    
    # Plot bars for each file
    for i, file_name in enumerate(file_names):
        if file_name in data:
            x_positions = np.arange(num_types) + (i - num_files/2 + 0.5) * bar_width
            values = [data[file_name][ordered_node_types[node_type]] for node_type in node_types]
            ax.bar(x_positions, values, width=bar_width, label=file_name)
    
    # Set labels and title
    ax.set_xlabel('Node Type')
    
    if data_key == "node_type_counts":
        if normalize:
            ax.set_ylabel('Proportion')
            ax.set_title('Node Type Proportions Comparison Across Models')
            output_file = output_dir / "node_type_proportions_comparison.png"
        else:
            ax.set_ylabel('Count')
            ax.set_title('Node Type Counts Comparison Across Models')
            output_file = output_dir / "node_type_counts_comparison.png"
    elif data_key == "node_avg_prob_sum":
        ax.set_ylabel('Average Probability')
        ax.set_title('Node Average Probabilities Comparison Across Models')
        output_file = output_dir / "node_avg_prob_comparison.png"
    else:
        ax.set_ylabel('Value')
        ax.set_title(f'{data_key} Comparison Across Models')
        output_file = output_dir / f"{data_key}_comparison.png"
        
    ax.set_xticks(np.arange(num_types))
    ax.set_xticklabels(node_types, rotation=45, ha='right')
    ax.legend()
    plt.tight_layout()
    
    # Save figure
    plt.savefig(output_file, dpi=300)
    plt.close()
    
    return str(output_file)


def create_node_prob_by_model_chart(file_paths: List[str], all_vectors_matrices: List[Dict[str, np.ndarray]], 
                                ordered_node_types: OrderedDict, output_dir: Path, 
                                normalize: bool = False, data_key: str = "node_avg_prob_sum") -> str:
    """
    Create a bar chart comparing node metrics across all models,
    with models on x-axis and node types as different bars for each model.
    
    Args:
        file_paths: List of analysis file paths
        all_vectors_matrices: List of vectors and matrices for each file
        ordered_node_types: OrderedDict of node types
        output_dir: Directory to save the chart
        normalize: Whether to normalize counts as proportions
        data_key: The key for the data to plot (node_type_counts or node_avg_prob_sum)
        
    Returns:
        Path to the saved chart file
    """
    # Extract file names for x-axis
    file_names = [get_model_name(path) for path in file_paths]
    
    # Collect data for each file and node type
    data = {}
    for node_type in ordered_node_types.keys():
        data[node_type] = []
    
    # For each file, collect values for each node type
    for i, (file_name, vectors_matrices) in enumerate(zip(file_names, all_vectors_matrices)):
        if data_key in vectors_matrices:
            values = vectors_matrices[data_key].copy()  # Make a copy to avoid modifying original data
            
            # Normalize if requested (only for counts, not for probabilities)
            if normalize and data_key == "node_type_counts":
                total = np.sum(values)
                if total > 0:
                    values = values / total
            
            # Add values to respective node type lists
            for node_type, idx in ordered_node_types.items():
                data[node_type].append(values[idx])
        else:
            # If data is missing, add zeros
            for node_type in ordered_node_types.keys():
                data[node_type].append(0)
    
    # Create plot
    fig = plt.figure(figsize=(14, 8))
    ax = fig.add_subplot(111)
    
    # Set up bar positions
    num_models = len(file_names)
    num_types = len(ordered_node_types)
    bar_width = 0.8 / num_types
    
    # Plot bars for each node type
    for i, (node_type, values) in enumerate(data.items()):
        x_positions = np.arange(num_models) + (i - num_types/2 + 0.5) * bar_width
        ax.bar(x_positions, values, width=bar_width, label=node_type)
    
    # Set labels and title
    ax.set_xlabel('Model')
    
    if data_key == "node_type_counts":
        if normalize:
            ax.set_ylabel('Proportion')
            ax.set_title('Node Type Proportions by Model')
            output_file = output_dir / "node_type_proportions_by_model_comparison.png"
        else:
            ax.set_ylabel('Count')
            ax.set_title('Node Type Counts by Model')
            output_file = output_dir / "node_type_counts_by_model_comparison.png"
    elif data_key == "node_avg_prob_sum":
        ax.set_ylabel('Average Probability')
        ax.set_title('Node Average Probabilities by Model')
        output_file = output_dir / "node_avg_prob_by_model_comparison.png"
    else:
        ax.set_ylabel('Value')
        ax.set_title(f'{data_key} by Model')
        output_file = output_dir / f"{data_key}_by_model_comparison.png"
    
    ax.set_xticks(np.arange(num_models))
    ax.set_xticklabels(file_names, rotation=45, ha='right')
    ax.legend(title='Node Type')
    plt.tight_layout()
    
    # Save figure
    plt.savefig(output_file, dpi=300)
    plt.close()
    
    return str(output_file)


def save_metrics_to_excel(file_paths: list, all_vectors_matrices: list, ordered_node_types: OrderedDict, output_dir: Path):
    """
    将每个模型的所有metric保存到xlsx文件，每个sheet一个模型。
    1D metric（如Node Type Counts）用单行表格，2D metric（如矩阵）用表格。
    各metric之间用空行隔开，上方注明metric名字。
    """
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    wb = Workbook()
    node_types = list(ordered_node_types.keys())
    metric_order = [
        ("node_type_counts", "Node Type Counts"),
        ("node_avg_prob_sum", "Node Average Probabilities"),
        ("dependency", "Dependency Matrix"),
        ("source_confidence", "Source Confidence Matrix"),
        ("target_confidence", "Target Confidence Matrix"),
        ("confidence_difference", "Source-Target Confidence Difference Matrix"),
        ("transition_counts", "Transition Counts Matrix")
    ]
    
    for i, (file_path, vectors_matrices) in enumerate(zip(file_paths, all_vectors_matrices)):
        model_name = get_model_name(file_path)
        if i == 0:
            ws = wb.active
            ws.title = model_name
        else:
            ws = wb.create_sheet(title=model_name)
        row_cursor = 1
        for metric_key, metric_name in metric_order:
            if metric_key in vectors_matrices:
                ws.cell(row=row_cursor, column=1, value=metric_name)
                row_cursor += 1
                arr = vectors_matrices[metric_key]
                if arr.ndim == 1:
                    # 1D metric: 单行表格
                    ws.append(["Node Type"] + node_types)
                    ws.append(["Value"] + [float(f"{v:.4f}") for v in arr])
                    row_cursor += 2
                elif arr.ndim == 2:
                    # 2D metric: 矩阵表格
                    ws.append(["/"] + node_types)
                    for idx, row in enumerate(arr):
                        ws.append([node_types[idx]] + [float(f"{v:.4f}") for v in row])
                    row_cursor += arr.shape[0] + 1
                row_cursor += 1  # 空行隔开
    # 保存
    xlsx_path = output_dir / "cpg_metrics.xlsx"
    wb.save(xlsx_path)
    print(f"Excel metrics saved to: {xlsx_path}")
    return str(xlsx_path)


def save_detailed_analysis(file_paths: List[str], all_vectors_matrices: List[Dict[str, np.ndarray]], similarities: Dict[str, Dict[Tuple[str, str], float]], ordered_node_types: OrderedDict, compare_similarity: bool = True) -> str:
    """
    保存详细分析结果到文本文件
    
    Args:
        file_paths: 分析的文件路径列表
        all_vectors_matrices: 所有文件的向量和矩阵数据
        similarities: 相似度计算结果
        ordered_node_types: 有序的节点类型字典
        compare_similarity: 是否保存相似度比较信息
        
    Returns:
        保存的文件路径
    """
    # 创建输出目录
    script_dir = Path(__file__).parent
    output_dir = script_dir / "cpg_detailed_analysis"
    output_dir.mkdir(exist_ok=True)
    
    # 创建输出文件
    output_file = output_dir / f"cpg_analysis.txt"
    
    # 创建节点类型计数比较图 (按节点类型分组)
    chart_file = create_node_type_counts_chart(file_paths, all_vectors_matrices, ordered_node_types, output_dir, 
                                              normalize=False, data_key="node_type_counts")
    
    # 创建节点类型比例比较图 (按节点类型分组)
    prop_chart_file = create_node_type_counts_chart(file_paths, all_vectors_matrices, ordered_node_types, output_dir, 
                                                  normalize=True, data_key="node_type_counts")
    
    # 创建节点平均概率比较图 (按节点类型分组)
    prob_chart_file = create_node_type_counts_chart(file_paths, all_vectors_matrices, ordered_node_types, output_dir,
                                                  normalize=False, data_key="node_avg_prob_sum")
    
    # 创建节点平均概率比较图 (按模型分组)
    prob_by_model_chart_file = create_node_prob_by_model_chart(file_paths, all_vectors_matrices, ordered_node_types, output_dir,
                                                             normalize=False, data_key="node_avg_prob_sum")
    
    # 创建节点类型计数比较图 (按模型分组)
    counts_by_model_chart_file = create_node_prob_by_model_chart(file_paths, all_vectors_matrices, ordered_node_types, output_dir,
                                                               normalize=False, data_key="node_type_counts")
    
    # 创建节点类型比例比较图 (按模型分组)
    prop_by_model_chart_file = create_node_prob_by_model_chart(file_paths, all_vectors_matrices, ordered_node_types, output_dir,
                                                             normalize=True, data_key="node_type_counts")
    
    # 打印调试信息
    print("\nAvailable data for analysis:")
    if all_vectors_matrices and len(all_vectors_matrices) > 0:
        first_file = all_vectors_matrices[0]
        print(f"Keys in first file: {list(first_file.keys())}")
    else:
        print("No vector matrices data available")
    
    with open(output_file, "w") as f:
        # 写入标题
        f.write("=" * 100 + "\n")
        f.write("LLM Cognitive Process Graph Analysis - Detailed Report\n")
        f.write(f"Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 100 + "\n\n")
        
        # 写入文件列表
        f.write("Analyzed Files:\n")
        f.write("-" * 80 + "\n")
        for i, filepath in enumerate(file_paths):
            f.write(f"{i+1}. {filepath}\n")
        f.write("\n\n")
        
        # 添加图表信息
        f.write("Visualization:\n")
        f.write("-" * 80 + "\n")
        f.write("按节点类型分组的图表:\n")
        if chart_file:
            f.write(f"  Node Type Counts Comparison Chart: {Path(chart_file).name}\n")
        if prop_chart_file:
            f.write(f"  Node Type Proportions Comparison Chart: {Path(prop_chart_file).name}\n")
        if prob_chart_file:
            f.write(f"  Node Average Probabilities Comparison Chart: {Path(prob_chart_file).name}\n")
            
        f.write("\n按模型分组的图表:\n")
        if counts_by_model_chart_file:
            f.write(f"  Node Type Counts by Model Chart: {Path(counts_by_model_chart_file).name}\n")
        if prop_by_model_chart_file:
            f.write(f"  Node Type Proportions by Model Chart: {Path(prop_by_model_chart_file).name}\n")
        if prob_by_model_chart_file:
            f.write(f"  Node Average Probabilities by Model Chart: {Path(prob_by_model_chart_file).name}\n")
        f.write("\n\n")
        
        # 写入节点类型
        f.write("Node Types (in alphabetical order):\n")
        f.write("-" * 50 + "\n")
        for i, node_type in enumerate(ordered_node_types.keys()):
            f.write(f"{i+1}. {node_type}\n")
        f.write("\n\n")
        
        # 写入每个文件的数据
        f.write("Individual File Analysis:\n")
        f.write("=" * 100 + "\n\n")
        
        for i, (filepath, vectors_matrices) in enumerate(zip(file_paths, all_vectors_matrices)):
            f.write(f"File {i+1}: {filepath}\n")
            f.write("-" * 80 + "\n\n")
            
            # 添加可用的数据键列表
            f.write(f"Available data keys: {list(vectors_matrices.keys())}\n\n")
            
            # 节点类型计数向量
            if "node_type_counts" in vectors_matrices:
                f.write("Node Type Counts:\n")
                for node_type, idx in ordered_node_types.items():
                    count = vectors_matrices["node_type_counts"][idx]
                    f.write(f"  {node_type}: {count:.4f}\n")
                f.write("\n")
            
            # 节点平均概率
            if "node_avg_prob_sum" in vectors_matrices:
                f.write("Node Average Probabilities:\n")
                for node_type, idx in ordered_node_types.items():
                    prob = vectors_matrices["node_avg_prob_sum"][idx]
                    f.write(f"  {node_type}: {prob:.4f}\n")
                f.write("\n")
            
            # 依赖矩阵
            if "dependency" in vectors_matrices:
                f.write("Dependency Matrix:\n")
                f.write("  " + " ".join(f"{node_type[:10]:>10}" for node_type in ordered_node_types.keys()) + "\n")
                
                for from_type, from_idx in ordered_node_types.items():
                    row_values = [f"{vectors_matrices['dependency'][from_idx, to_idx]:.2f}" for to_idx in range(len(ordered_node_types))]
                    f.write(f"{from_type[:10]:>10} " + " ".join(f"{val:>10}" for val in row_values) + "\n")
                f.write("\n")
            
            # 源置信度矩阵
            if "source_confidence" in vectors_matrices:
                f.write("Source Confidence Matrix:\n")
                f.write("  " + " ".join(f"{node_type[:10]:>10}" for node_type in ordered_node_types.keys()) + "\n")
                
                for from_type, from_idx in ordered_node_types.items():
                    row_values = [f"{vectors_matrices['source_confidence'][from_idx, to_idx]:.4f}" for to_idx in range(len(ordered_node_types))]
                    f.write(f"{from_type[:10]:>10} " + " ".join(f"{val:>10}" for val in row_values) + "\n")
                f.write("\n")
            else:
                f.write("Source Confidence Matrix: Not available in this data\n\n")
            
            # 目标置信度矩阵
            if "target_confidence" in vectors_matrices:
                f.write("Target Confidence Matrix:\n")
                f.write("  " + " ".join(f"{node_type[:10]:>10}" for node_type in ordered_node_types.keys()) + "\n")
                
                for from_type, from_idx in ordered_node_types.items():
                    row_values = [f"{vectors_matrices['target_confidence'][from_idx, to_idx]:.4f}" for to_idx in range(len(ordered_node_types))]
                    f.write(f"{from_type[:10]:>10} " + " ".join(f"{val:>10}" for val in row_values) + "\n")
                f.write("\n")
            else:
                f.write("Target Confidence Matrix: Not available in this data\n\n")
            
            # 源到目标置信度差异矩阵
            if "confidence_difference" in vectors_matrices:
                f.write("Source-Target Confidence Difference Matrix (Source - Target):\n")
                f.write("  " + " ".join(f"{node_type[:10]:>10}" for node_type in ordered_node_types.keys()) + "\n")
                
                for from_type, from_idx in ordered_node_types.items():
                    row_values = [f"{vectors_matrices['confidence_difference'][from_idx, to_idx]:.4f}" for to_idx in range(len(ordered_node_types))]
                    f.write(f"{from_type[:10]:>10} " + " ".join(f"{val:>10}" for val in row_values) + "\n")
                f.write("\n")
            else:
                f.write("Confidence Difference Matrix: Not available in this data\n\n")
            
            # 转换计数矩阵
            if "transition_counts" in vectors_matrices:
                f.write("Transition Counts Matrix:\n")
                f.write("  " + " ".join(f"{node_type[:10]:>10}" for node_type in ordered_node_types.keys()) + "\n")
                
                for from_type, from_idx in ordered_node_types.items():
                    row_values = [f"{vectors_matrices['transition_counts'][from_idx, to_idx]:.0f}" for to_idx in range(len(ordered_node_types))]
                    f.write(f"{from_type[:10]:>10} " + " ".join(f"{val:>10}" for val in row_values) + "\n")
                f.write("\n")
            else:
                f.write("Transition Counts Matrix: Not available in this data\n\n")
            
            f.write("\n" + "=" * 100 + "\n\n")
        
        # 写入相似度比较，只在compare_similarity为True时执行
        if compare_similarity:
            f.write("Similarity Analysis:\n")
            f.write("=" * 100 + "\n\n")
            
            for metric, pairs in similarities.items():
                if not pairs:
                    continue
                    
                f.write(f"{metric} Similarities:\n")
                f.write("-" * 80 + "\n")
                
                for (file1, file2), similarity in pairs.items():
                    # 提取文件名部分，避免路径太长导致输出难以阅读
                    file1_name = Path(file1).name
                    file2_name = Path(file2).name
                    f.write(f"  {file1_name} vs {file2_name}: {similarity:.4f}\n")
                    f.write(f"    Full paths:\n")
                    f.write(f"    - {file1}\n")
                    f.write(f"    - {file2}\n\n")
                
                # 计算平均相似度
                avg_similarity = sum(pairs.values()) / len(pairs)
                f.write(f"\n  Average {metric} similarity: {avg_similarity:.4f}\n\n")
        
        f.write("\n" + "=" * 100 + "\n")
        f.write("End of Report\n")
    
    # 保存所有metric到excel
    save_metrics_to_excel(file_paths, all_vectors_matrices, ordered_node_types, output_dir)
    
    return str(output_file)


def extract_and_process_metadata(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Extracts and processes metadata from analysis file, with special handling for transitions data
    
    Args:
        data: The loaded JSON data
        
    Returns:
        Processed metadata with additional metrics if needed
    """
    if "metadata" not in data:
        return {}
        
    metadata = data["metadata"]
    metadata["average_summary_metrics"]["all_confidence_transitions"] = metadata["all_confidence_transitions"]
    
    return metadata


def compare_analysis_files(file_paths: List[str]):
    """
    Compare multiple LLM analysis files and calculate similarities.
    
    Args:
        file_paths: List of paths to LLM analysis JSON files
        
    Returns:
        Dictionary mapping metric names to dictionaries of file pairs to similarity scores
    """
    # Load all files
    analysis_data = []
    file_paths_valid = []
    
    for file_path in file_paths:
        path = Path(file_path)
        if not path.exists():
            print(f"Warning: File not found: {file_path}")
            continue
            
        try:
            data = load_analysis_file(file_path)
            # Process metadata to ensure transitions data is available
            processed_data = {"metadata": extract_and_process_metadata(data)}
            
            if not processed_data["metadata"]:
                print(f"Warning: No valid metadata found in file: {file_path}")
                continue
                
            file_paths_valid.append(str(path))
            analysis_data.append(processed_data)
            print(f"Successfully loaded and processed: {path.name}")
        except Exception as e:
            print(f"Error processing file {file_path}: {e}")
            continue
    
    if len(analysis_data) < 2:
        print("Error: At least two valid analysis files are required for comparison")
        return {}, [], [], OrderedDict()
    
    # Collect all node types across all files
    all_node_types = set()
    for data in analysis_data:
        metadata = data.get("metadata", {})
        metrics = metadata.get("average_summary_metrics", {})
        
        for key in ["node_type_counts", "node_avg_prob_sum", "dependency"]:
            if key in metrics:
                all_node_types.update(metrics[key].keys())
    
    print(f"Found {len(all_node_types)} node types across all files: {sorted(list(all_node_types))}")
    
    # Create an ordered dictionary of all node types
    all_ordered_node_types = OrderedDict([(node_type, i) for i, node_type in enumerate(sorted(all_node_types))])
    
    # Extract vectors and matrices for each file
    all_vectors_matrices = []
    for data in analysis_data:
        metadata = data.get("metadata", {})
        metrics = metadata.get("average_summary_metrics", {})
        vectors_matrices = create_vectors_and_matrices(all_ordered_node_types, metrics)
        all_vectors_matrices.append(vectors_matrices)
    
    # Calculate similarities between each pair of files
    similarities = {
        "node_type_counts": {},
        "node_avg_prob_sum": {},
        "dependency": {},
        "source_confidence": {},
        "target_confidence": {},
        "confidence_difference": {},
        "transition_counts": {}
    }
    
    for i, j in itertools.combinations(range(len(file_paths_valid)), 2):
        file_pair = (file_paths_valid[i], file_paths_valid[j])
        vm_i = all_vectors_matrices[i]
        vm_j = all_vectors_matrices[j]
        
        # Compare node_type_counts
        if "node_type_counts" in vm_i and "node_type_counts" in vm_j:
            similarities["node_type_counts"][file_pair] = compute_cosine_similarity(
                vm_i["node_type_counts"], vm_j["node_type_counts"]
            )
        
        # Compare node_avg_prob_sum
        if "node_avg_prob_sum" in vm_i and "node_avg_prob_sum" in vm_j:
            similarities["node_avg_prob_sum"][file_pair] = compute_cosine_similarity(
                vm_i["node_avg_prob_sum"], vm_j["node_avg_prob_sum"]
            )
        
        # Compare dependency matrices
        if "dependency" in vm_i and "dependency" in vm_j:
            similarities["dependency"][file_pair] = compute_matrix_similarity(
                vm_i["dependency"], vm_j["dependency"]
            )
        
        # Compare source_confidence matrices
        if "source_confidence" in vm_i and "source_confidence" in vm_j:
            similarities["source_confidence"][file_pair] = compute_matrix_similarity(
                vm_i["source_confidence"], vm_j["source_confidence"]
            )
        
        # Compare target_confidence matrices
        if "target_confidence" in vm_i and "target_confidence" in vm_j:
            similarities["target_confidence"][file_pair] = compute_matrix_similarity(
                vm_i["target_confidence"], vm_j["target_confidence"]
            )
            
        # Compare confidence_difference matrices
        if "confidence_difference" in vm_i and "confidence_difference" in vm_j:
            similarities["confidence_difference"][file_pair] = compute_matrix_similarity(
                vm_i["confidence_difference"], vm_j["confidence_difference"]
            )
            
        # Compare transition_counts matrices
        if "transition_counts" in vm_i and "transition_counts" in vm_j:
            similarities["transition_counts"][file_pair] = compute_matrix_similarity(
                vm_i["transition_counts"], vm_j["transition_counts"]
            )
    
    return similarities, file_paths_valid, all_vectors_matrices, all_ordered_node_types


def get_default_files():
    """
    递归查找当前目录及其子目录下所有以_logical_graph_llm_analysis.json结尾的文件
    """
#     script_dir = Path(__file__).parent
    script_dir ="/staging/szhang967/results_temperature_0"
    result = []
    
    for root, dirs, files in os.walk(script_dir):
        for file in files:
            if file.endswith('cognitive_process_graph_logical_graph_llm_analysis.json'):
                result.append(os.path.join(root, file))
    
    return result


def main():
    parser = argparse.ArgumentParser(description='Compare LLM analysis files')
    parser.add_argument('--files', type=str, nargs='+', required=False,
                        default=get_default_files(),
                        help='Paths to LLM analysis JSON files (default: all *_logical_graph_llm_analysis.json files in current directory)')
    parser.add_argument('--output-txt', action='store_true',
                        help='Generate detailed analysis text file')
    parser.add_argument('--compare-similarity', action='store_true', default=False,
                        help='Include similarity comparison information in the output text file')
    
    args = parser.parse_args()
    
    # Check if any files were provided
    if not args.files:
        print("Error: No input files provided")
        return
    
    # Compare the files
    similarities, file_paths_valid, all_vectors_matrices, all_ordered_node_types = compare_analysis_files(args.files)
    
    # Print the results
    print("\nSimilarity Analysis Results:")
    print("============================")
    
    for metric, pairs in similarities.items():
        print(f"\n{metric} Similarities:")
        print("-" * (len(metric) + 13))
        
        for (file1, file2), similarity in pairs.items():
            # 使用Path.name提取文件名，但保留完整路径用于详细输出
            file1_name = Path(file1).name
            file2_name = Path(file2).name
            print(f"{file1_name} vs {file2_name}: {similarity:.4f}")
    
    # Print summary
    print("\nSummary:")
    print("========")
    
    for metric in similarities.keys():
        if similarities[metric]:
            avg_similarity = sum(similarities[metric].values()) / len(similarities[metric])
            print(f"Average {metric} similarity: {avg_similarity:.4f}")
    
    # 保存详细分析到文本文件 (默认始终生成)
    if True:
        # 直接使用从compare_analysis_files返回的数据
        output_file = save_detailed_analysis(file_paths_valid, all_vectors_matrices, similarities, all_ordered_node_types, args.compare_similarity)
        print(f"\nDetailed analysis saved to: {output_file}")


if __name__ == "__main__":
    main() 